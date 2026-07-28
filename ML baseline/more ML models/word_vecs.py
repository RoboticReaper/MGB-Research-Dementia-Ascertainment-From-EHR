import sys

from gensim.models import Word2Vec
from nltk.tokenize import word_tokenize, sent_tokenize
from openpyxl import load_workbook
import psycopg2

from utils import clinic_text_processing
from ds_prep import find_sentences


def load_text_data_xlsx(input_report, X_id):
    r''' load the xlsx data
        Args:
            input_report (string): input xlsx file directory
            X_id (string): the column id of input X ('A', 'B',...)
        Return:
            Text data X for embedding input.
    '''
    print("Start loading xlsx file from %s."%input_report)
    report = load_workbook(input_report)
    active_workbook = report.active
    print("\tworkbook: %s"%(report.sheetnames))
    ws = report[report.sheetnames[0]]
    instance_num = len(ws['A']) - 1
    X = []
    for idx in range(1, instance_num+1):
        the_x = ws[X_id+str(idx)].value
        the_x = clinic_text_processing(the_x)
        X.append(the_x)

    instance_num = len(X)
    print("\tInstance Num: %s"%instance_num)
    print("File loaded.")
    return X


def load_text_data(input_file):
    with open(input_file, 'r', encoding='utf-8') as f:
        notes = [split_note(line) for line in f]

    return notes


def load_text_data_db(db, schema, table, username, pword, note_column_idx):
    conn = psycopg2.connect(host='voice', user=username, password=pword, database=db)
    cur = conn.cursor()
    cur.execute(f'SELECT * FROM {schema}.{table}')
    rows = cur.fetchall()
    notes = [note_data[note_column_idx] for note_data in rows]

    return notes


def split_note(note_data):
    note_split = note_data.split('\t')
    note = ' '.join(note_split[1:])

    return note


if __name__ == '__main__':
    print('loading notes...')
    # input_file = r"../data/scd_annotation_sample.xlsx"
    # notes = load_text_data_xlsx(input_file, 'F')
    # input_file = r"../data/embedding_test.txt"
    # notes = load_text_data(input_file)
    notes = load_text_data_db('palliative', 'scd', 'mci_cohort_icd_jan2017feb2020_epic_2010feb2020_merged',
                              username=sys.argv[1], pword=sys.argv[2], note_column_idx=4)
    print(f'{len(notes)} notes loaded')

    # # if creating embeddings at the note level, just use word_tokenize
    # tokenized_notes = [word_tokenize(note) for note in x]

    # if creating embeddings at the sentence level, need to split on sentences first
    print('tokenizing notes...')
    char_embeddings = False  # if creating character embeddings, set to True
    # sents_by_note = [sent_tokenize(note) for note in notes]
    # tokenized_sents = [word_tokenize(sent) for note_sents in sents_by_note for sent in note_sents]
    sents_by_note = [find_sentences(note, char_level=char_embeddings) for note in notes]
    # need to flatten from list of lists of lists to a list of lists (i.e. put all tokenized sentences into one list)
    tokenized_sents = [sent for note_sents in sents_by_note for sent in note_sents]
    print('notes tokenized')

    model = Word2Vec(tokenized_sents)

    # save the model
    model_file = 'embeddings/2017_2020_cohort/gensim_word2vec.model'
    model.save(model_file)
    print(f'embedding model saved to {model_file}')
    # save embeddings
    embedding_file = 'embeddings/2017_2020_cohort/gensim_word2vec.emb'
    model.wv.save_word2vec_format(embedding_file, binary=False)
    print(f'embeddings saved to {embedding_file}')

    # # load model
    # model = Word2Vec.load('dl_test_output/word2vec_loadable_format_test.emb')
