# -*- coding: utf-8 -*-
# @Author: Jie
# @Date:   2017-06-15 14:11:08
# @Last Modified by:   Jie Yang,     Contact: jieynlp@gmail.com
# @Last Modified time: 2020-02-19 10:34:00


import time
import sys
import random
import csv

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn import metrics
from sklearn.metrics import roc_auc_score
from sklearn.metrics import average_precision_score
import nltk

from NCRFpp.ncrf import NCRF
from file_io import *
from utils import clinic_text_list_processing, clinic_text_processing, data_split_train_dev_test, filter_duplicate
from metric_visual import *
import stat_util

# np.set_printoptions(threshold=np.nan)


def write_results(out_file, input_list, name_list, auc=None):
    print("Write results:...")
    if out_file.endswith('.csv'):
        auc_out_file = out_file.replace('.csv', '.txt')
        fout = open(auc_out_file, 'w')
        if auc != None:
            fout.write("AUC: " + str(auc) + "\n")
        fout.close()

        pr_out_file = out_file[:-4] + '.proutput.csv'
        with open(pr_out_file, 'w', newline='') as f:
            # write precision/recall results
            output_writer = csv.writer(f)
            output_writer.writerow(name_list)
            for instance_results in zip(*input_list[:3]):
                output_writer.writerow(instance_results)
        with open(out_file, 'w', newline='') as f:
            # write roc results
            output_writer = csv.writer(f)
            output_writer.writerow(name_list)
            for instance_results in zip(*input_list[3:]):
                output_writer.writerow(instance_results)
    else:
        fout = open(out_file, 'w')
        if auc != None:
            fout.write("AUC: " + str(auc) + "\n")

        assert (len(input_list) == len(name_list))
        for name, the_list in zip(name_list, input_list):
            fout.write(name + ": " + " ".join([str(a) for a in the_list.tolist()]) + "\n")
        fout.close()
    print("Results are written into file:", out_file)


def load_text_data_xlsx(input_report, X_id, Y_id, id_id='A', shuffle=False, filter_duplicacte_instances=True, include_title=True):
    r''' load the xlsx train data: 9K_reports_20181228.xlsx
        Args:
            input_report (string): input xlsx file directory
            X_id (string): the column id of input X ('A', 'B',...)
            Y_id (string): the column id of label Y ('A', 'B',...)
            id_id (string): the column id of note IDs
        Return:
            Descriptions, ADE, ADR, HSR, MRN: all lists, corresponding information, text is preprocessed.
    '''
    print("Start loading xlsx file from %s."%input_report)
    report = load_workbook(input_report)
    active_workbook = report.active
    print("\tworkbook: %s"%(report.sheetnames))
    ws = report[report.sheetnames[0]]
    instance_num = len(ws['A']) - 1
    X = []
    Y = []
    id_list = []
    for idx in range(1, instance_num+2):
        the_x = ws[X_id+str(idx)].value
        the_y = str(ws[Y_id+str(idx)].value)
        the_id = ws[id_id+str(idx)].value
        if the_y in ['0','1']:
            the_x = clinic_text_processing(the_x)
            X.append(the_x)
            Y.append(int(the_y))
            id_list.append(the_id)

    if filter_duplicacte_instances:
        X,Y,id_list = filter_duplicate([X,Y,id_list])
        # X,Y = filter_duplicate([X,Y])
    instance_num = len(X)
    print("\tInstance Num: %s"%instance_num)
    if shuffle:
        combined = list(zip(X,Y,id_list))
        # combined = list(zip(X,Y))
        random.shuffle(combined)
        X[:],Y[:],id_list[:] = zip(*combined)
        # X[:],Y[:] = zip(*combined)
        print("\tInstances shuffled.")
    print("File loaded.")
    return X, Y, id_list
    # return X, Y


def nfold_hsr_arg(input_file, nfold=5, word_feature="FF", use_char=False, lr=0.2, cutoff=5):
    r''' perform cross-fold validation
        Args:
            x (list): size[num_instance], list of strings (sentences/documents)
            y (list): size[num_instance], list of labels
            nfold (int): number of folds
        Return:
    '''
    ## select the text and label from excel
    x, y, id_list = load_text_data_xlsx(input_file, 'B', 'C', shuffle=True)
    # x, y, id_list = load_text_data_xlsx(input_file, 'F', 'L', shuffle=True, filter_duplicacte_instances=False)

    print("Run nfold experiments for %s, nfold=%s" % ("HSR", nfold))

    test_decode_list = []
    test_gold_list = []
    test_attn_list = []
    test_id_list = []
    for idx in range(nfold):
        start_time = time.time()
        print("Proceesing: %s/%s.............." % (idx, nfold))
        ## notice the positive label id may vary in different fold, need fix the positive label id.
        train_x, dev_x, test_x = data_split_train_dev_test(x, nfold, idx)
        train_y, dev_y, test_y = data_split_train_dev_test(y, nfold, idx)
        train_ids, dev_ids, test_ids = data_split_train_dev_test(id_list, nfold, idx)
        test_id_list.extend(test_ids)
        train_list = convert_to_ncrf_list(train_x, train_y)
        dev_list = convert_to_ncrf_list(dev_x, dev_y)
        test_list = convert_to_ncrf_list(test_x, test_y)
        config_file = "NCRFpp/2019_cohort.config"
        ncrf = NCRF()
        ncrf.read_data_config_file(config_file)
        ncrf.data.word_feature_extractor = word_feature
        ncrf.data.use_char = use_char
        # ncrf.data.char_emb_dir = "embeddings/2019_cohort/gensim_word2vec_char.emb"
        ncrf.data.word_emb_dir = "../../embeddings/aav_cohort/gensim_word2vec_aav.emb"
        #ncrf.data.word_emb_dir = "../../case_finding/code/embeddings/2017_2020_cohort/gensim_word2vec.emb"
        # ncrf.data.word_emb_dir = "embeddings/2017_2020_cohort/gensim_word2vec.emb"
        # ncrf.data.word_emb_dir = "embeddings/2019_cohort/gensim_word2vec.emb"
        # ncrf.data.word_emb_dir = "MGH.lower.emb"
        ncrf.data.word_cutoff = cutoff
        # ncrf.data.optimizer = 'adagrad'
        ncrf.data.optimizer = "sgd"
        # ncrf.data.HP_batch_size = 5
        ncrf.data.HP_lr = lr
        ncrf.data.HP_dropout = 0.5
        # ncrf.data.HP_dropout = 0.3
        ncrf.data.char_feature_extractor = "CNN"
        ncrf.data.words2sent_representation = "ATTENTION"

        ncrf.data.HP_iteration = 30

        # model_name = ncrf.data.word_feature_extractor + "." + ncrf.data.words2sent_representation + "." + str(
        #     ncrf.data.use_char) + ncrf.data.char_feature_extractor + "." + ncrf.data.word_emb_dir.replace('/', '.') + ".opt" + ncrf.data.optimizer + ".wcut" + str(
        #     ncrf.data.word_cutoff) + ".lr" + str(ncrf.data.HP_lr) + ".h" + str(ncrf.data.HP_hidden_dim)
        # model_name = "HSR." + model_name
        # model_name = model_name + ncrf.data.optimizer
        model_name = 'test_final_v2'
        print("Model Name:", model_name)
        ncrf.data.model_dir = "dl_output/aav_cohort/" + model_name + ".model"
        # ncrf.data.model_dir = "dl_output/2019_cohort/" + model_name + ".model"
        ncrf.initialization([train_list, dev_list, test_list])
        ncrf.generate_instances_from_list(train_list, 'train')
        ncrf.generate_instances_from_list(dev_list, 'dev')
        test_Ids = ncrf.generate_instances_from_list(test_list, 'test')
        ncrf.data.show_data_summary()
        ncrf.train()
        target_prob, attention_weights = ncrf.decode_prob_and_attention_weights(ncrf.data.test_Ids)
        # target_prob = ncrf.decode_prob(ncrf.data.test_Ids)
        test_decode_list.append(np.asarray(target_prob))
        test_gold_list.append(np.asarray(test_y))
        # test_attn_list.append(np.asarray(attention_weights))
        test_text_tokens = [text_data[0] for text_data in ncrf.data.test_texts]
        with open(f"dl_output/attn_test/attn_weights_{idx}.txt", 'w', encoding='utf-8') as f:
            for test_idx, inst_tokens in enumerate(test_text_tokens):
                matched_attn_weights = list(zip(inst_tokens, attention_weights[test_idx]))
                for matched_pair in matched_attn_weights:
                    f.write(f'{str(matched_pair)}\n')
                f.write('\n\n')

        with open("dl_output/attn_test/attn_weights_final_v2.csv", 'a', newline='') as f:
            data_writer = csv.writer(f)
            if idx == 0:
                data_writer.writerow(['id', 'gold_label', 'prob', 'text', 'attn', 'attn_spillover'])

            for test_idx, inst_tokens in enumerate(test_text_tokens):
                matched_attn_weights = list(zip(inst_tokens, attention_weights[test_idx]))
                formatted_attn_weights = [f'{token}|{attn_weight}' for token, attn_weight in matched_attn_weights]
                attn_weight_string = ' '.join(formatted_attn_weights)
                data_writer.writerow([test_ids[test_idx], test_y[test_idx], target_prob[test_idx][-1], test_x[test_idx],
                                      attn_weight_string[:32700], attn_weight_string[32700:]])

    test_all_decode = np.concatenate(test_decode_list, axis=0)
    test_all_gold = np.concatenate(test_gold_list, axis=0)

    target_id = ncrf.data.label_alphabet.get_index('1')
    test_all_target_decode = test_all_decode[:, target_id]
    test_all_target_decode = np.nan_to_num(test_all_target_decode)
    target_label = 1

    precision, recall, pr_thresholds = metrics.precision_recall_curve(test_all_gold, test_all_target_decode,
                                                                      target_label)
    auprc = metrics.auc(recall, precision)
    fpr, tpr, roc_thresholds = metrics.roc_curve(test_all_gold, test_all_target_decode, target_label)
    auc = metrics.auc(fpr, tpr)

    plot_precision_recall(precision, recall, model_name, "dl_output/aav_cohort/" + model_name + ".prc.jpg")
    plot_roc(fpr, tpr, auc, model_name, "dl_output/aav_cohort/" + model_name + ".roc.jpg")
    write_results("dl_output/aav_cohort/" + model_name + ".result.csv", [pr_thresholds, precision, recall, roc_thresholds, fpr, tpr],
                  ["P/R Thresholds", "Precision", "Recall", "ROC Thresholds", "FPR", "TPR"], auc)
    print("auroc:", auc)
    print("auprc:", auprc)

    # auc 95% confidence interval
    score, ci_lower, ci_upper, scores = stat_util.score_ci(
        test_all_gold, test_all_target_decode, score_fun=roc_auc_score
    )
    print('\n95% confidence interval:')
    print(f'auroc: {score}')
    print(f'lower: {ci_lower}')
    print(f'upper: {ci_upper}\n')

    score, ci_lower, ci_upper, scores = stat_util.score_ci(
        test_all_gold, test_all_target_decode, score_fun=average_precision_score
    )
    print('\n95% confidence interval:')
    print(f'auprc: {score}')
    print(f'lower: {ci_lower}')
    print(f'upper: {ci_upper}\n')

    return test_all_target_decode, test_all_gold, test_id_list


def run_hsr_arg(train_file, test_file, word_feature="FF", use_char=False, lr=0.2, cutoff=5):
    r''' train model based on train data and predict on test data
        Args:
            x (list): size[num_instance], list of strings (sentences/documents)
            y (list): size[num_instance], list of labels
            nfold (int): number of nfold
        Return:
    '''
    ## select the text and label from excel
    #train_x, train_y, train_ids = load_text_data_xlsx(train_file, 'G', 'M', shuffle=True)
    #test_x, test_y, test_ids = load_text_data_xlsx(test_file, 'G', 'M', filter_duplicacte_instances=False)
    train_x, train_y, train_ids = load_text_data_xlsx(train_file, 'G', 'M', shuffle=True)
    test_x, test_y, test_ids = load_text_data_xlsx(test_file, 'G', 'M', filter_duplicacte_instances=False)

    test_decode_list = []
    test_gold_list = []
    start_time = time.time()
    train_list = convert_to_ncrf_list(train_x, train_y)
    test_list = convert_to_ncrf_list(test_x, test_y)
    config_file = "NCRFpp/2019_cohort.config"
    ncrf = NCRF()
    ncrf.read_data_config_file(config_file)
    ncrf.data.word_feature_extractor = word_feature
    ncrf.data.use_char = use_char
    ncrf.data.word_emb_dir = "../../embeddings/aav_cohort/gensim_word2vec_aav.emb"  
    ##"../../case_finding/code/embeddings/2017_2020_cohort/gensim_word2vec.emb"
    ncrf.data.word_cutoff = cutoff
    ncrf.data.optimizer = "sgd"
    ncrf.data.HP_lr = lr
    ncrf.data.HP_dropout = 0.5
    ncrf.data.char_feature_extractor = "CNN"
    ncrf.data.words2sent_representation = "ATTENTION"
    ncrf.data.HP_iteration = 30

    # model_name = ncrf.data.word_feature_extractor + "." + ncrf.data.words2sent_representation + "." + str(
    #     ncrf.data.use_char) + ncrf.data.char_feature_extractor + "." + ncrf.data.word_emb_dir.replace('/', '.') + ".opt" + ncrf.data.optimizer + ".wcut" + str(
    #     ncrf.data.word_cutoff) + ".lr" + str(ncrf.data.HP_lr) + ".h" + str(ncrf.data.HP_hidden_dim)
    # model_name = "HSR." + model_name
    # model_name = model_name + ncrf.data.optimizer
    model_name = 'gen_final_corrected_testing'
    print("Model Name:", model_name)
    ncrf.data.model_dir = "dl_output/generalizability_output/updated_v2/" + model_name + ".model"
    ncrf.initialization([train_list, test_list, test_list])
    ncrf.generate_instances_from_list(train_list, 'train')
    ncrf.generate_instances_from_list(test_list, 'dev')
    test_Ids = ncrf.generate_instances_from_list(test_list, 'test')
    ncrf.data.show_data_summary()
    ncrf.train()
    target_prob, attention_weights = ncrf.decode_prob_and_attention_weights(ncrf.data.test_Ids)
    test_decode_list.append(np.asarray(target_prob))
    test_gold_list.append(np.asarray(test_y))
    test_text_tokens = [text_data[0] for text_data in ncrf.data.test_texts]
    with open(f"dl_output/generalizability_output/updated_v2/attn_weights.txt", 'w', encoding='utf-8') as f:
        for test_idx, inst_tokens in enumerate(test_text_tokens):
            matched_attn_weights = list(zip(inst_tokens, attention_weights[test_idx]))
            for matched_pair in matched_attn_weights:
                f.write(f'{str(matched_pair)}\n')
            f.write('\n\n')

    with open("dl_output/generalizability_output/updated_v2/attn_weights.csv", 'a', newline='') as f:
        data_writer = csv.writer(f)
        data_writer.writerow(['id', 'gold_label', 'prob', 'text', 'attn', 'attn_spillover'])

        for test_idx, inst_tokens in enumerate(test_text_tokens):
            matched_attn_weights = list(zip(inst_tokens, attention_weights[test_idx]))
            formatted_attn_weights = [f'{token}|{attn_weight}' for token, attn_weight in matched_attn_weights]
            attn_weight_string = ' '.join(formatted_attn_weights)
            data_writer.writerow(['x', test_y[test_idx], target_prob[test_idx][-1], test_x[test_idx],
                                  attn_weight_string[:32700], attn_weight_string[32700:]])

    test_all_decode = np.concatenate(test_decode_list, axis=0)
    test_all_gold = np.concatenate(test_gold_list, axis=0)

    target_id = ncrf.data.label_alphabet.get_index('1')
    test_all_target_decode = test_all_decode[:, target_id]
    test_all_target_decode = np.nan_to_num(test_all_target_decode)
    target_label = 1

    precision, recall, pr_thresholds = metrics.precision_recall_curve(test_all_gold, test_all_target_decode,
                                                                      target_label)
    auprc = metrics.auc(recall, precision)
    fpr, tpr, roc_thresholds = metrics.roc_curve(test_all_gold, test_all_target_decode, target_label)
    auc = metrics.auc(fpr, tpr)

    plot_precision_recall(precision, recall, model_name, "dl_output/generalizability_output/updated_v2/" + model_name + ".prc.jpg")
    plot_precision_recall(precision, recall, model_name, "dl_output/generalizability_output/updated_v2/" + model_name + ".prc.jpg")
    plot_roc(fpr, tpr, auc, model_name, "dl_output/generalizability_output/updated_v2/" + model_name + ".roc.jpg")
    write_results("dl_output/generalizability_output/updated_v2/" + model_name + ".result.csv", [pr_thresholds, precision, recall, roc_thresholds, fpr, tpr],
                  ["P/R Thresholds", "Precision", "Recall", "ROC Thresholds", "FPR", "TPR"], auc)
    print("auroc:", auc)
    print("auprc:", auprc)
    # fout = open("dl_output/2019_cohort/all_results.txt", 'a')
    # fout.write(model_name + ":" + str(auc) + "\n")
    # fout.close()

    # auc 95% confidence interval
    score, ci_lower, ci_upper, scores = stat_util.score_ci(
        test_all_gold, test_all_target_decode, score_fun=roc_auc_score
    )
    print('\n95% confidence interval:')
    print(f'auroc: {score}')
    print(f'lower: {ci_lower}')
    print(f'upper: {ci_upper}\n')

    score, ci_lower, ci_upper, scores = stat_util.score_ci(
        test_all_gold, test_all_target_decode, score_fun=average_precision_score
    )
    print('\n95% confidence interval:')
    print(f'auprc: {score}')
    print(f'lower: {ci_lower}')
    print(f'upper: {ci_upper}\n')

    return test_all_target_decode, test_all_gold, test_ids


def convert_to_ncrf_list(sent_list, label_list):
    word_list = []
    feature_list = []
    strlabel_list = []
    # informative_feature_list = load_informative_features('results/feature_selection/chi_square_+bigrams.txt')
    for sent, label in zip(sent_list, label_list):
        words = nltk.word_tokenize(sent)
        # word_list.append([term for term in words if term.lower() in informative_feature_list])
        word_list.append(words)
        strlabel_list.append(str(label))
        feature_list.append([])
    return [word_list, strlabel_list, feature_list]


def load_informative_features(input_file):
    feature_terms = set()
    with open(input_file, 'r', encoding='utf-8') as f:
        for line in f:
            feature_data = line.split('\t')
            p_value = float(feature_data[2])
            term = feature_data[0].lower()
            if p_value < 0.5:
                feature_terms.add(term)

    return feature_terms


def compute_metrics(prob_file):
    model_name = 'generalizability_v2'

    prob_df = pd.read_csv(prob_file)
    target_label = 1
    test_all_gold = np.array(prob_df['gold'])
    test_all_target_decode = np.array(prob_df['dl'])

    precision, recall, pr_thresholds = metrics.precision_recall_curve(test_all_gold, test_all_target_decode,
                                                                      target_label)
    auprc = metrics.auc(recall, precision)
    fpr, tpr, roc_thresholds = metrics.roc_curve(test_all_gold, test_all_target_decode, target_label)
    auc = metrics.auc(fpr, tpr)

    plot_precision_recall(precision, recall, model_name, "dl_output/generalizability_output/" + model_name + ".prc.jpg")
    plot_precision_recall(precision, recall, model_name, "dl_output/generalizability_output/" + model_name + ".prc.jpg")
    plot_roc(fpr, tpr, auc, model_name, "dl_output/generalizability_output/" + model_name + ".roc.jpg")
    write_results("dl_output/generalizability_output/" + model_name + ".result.csv", [pr_thresholds, precision, recall, roc_thresholds, fpr, tpr],
                  ["P/R Thresholds", "Precision", "Recall", "ROC Thresholds", "FPR", "TPR"], auc)

    print("auroc:", auc)
    print("auprc:", auprc)

    # auc 95% confidence interval
    score, ci_lower, ci_upper, scores = stat_util.score_ci(
        test_all_gold, test_all_target_decode, score_fun=roc_auc_score
    )
    print('\n95% confidence interval:')
    print(f'auroc: {score}')
    print(f'lower: {ci_lower}')
    print(f'upper: {ci_upper}\n')

    score, ci_lower, ci_upper, scores = stat_util.score_ci(
        test_all_gold, test_all_target_decode, score_fun=average_precision_score
    )
    print('\n95% confidence interval:')
    print(f'auprc: {score}')
    print(f'lower: {ci_lower}')
    print(f'upper: {ci_upper}\n')


def filter_duplicate_and_write(input_list, output_file, indent='\t', filter_id=0):
    ## remove duplicate instance based on input_list[0]
    item_num = len(input_list)
    original_size = len(input_list[filter_id])
    input_dict = {}
    output_list = [[] for _ in range(item_num)]
    for idx in range(original_size):
        if input_list[filter_id][idx] not in input_dict:
            input_dict[input_list[filter_id][idx]] = 1
            for idy in range(item_num):
                output_list[idy].append(input_list[idy][idx])
        else:
            continue
    new_size = len(output_list[filter_id])
    print(indent+"Filter duplicate data size: %s -> %s" %(original_size, new_size))

    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        filtered_result_writer = csv.writer(f)
        for row_data in zip(*output_list):
            filtered_result_writer.writerow(list(row_data))


if __name__ == '__main__':
    # ncrf = NCRF()
    # ncrf.load('dl_output/2017_2020_cohort/test_final.model.model')
    # exit(0)
    ## as our code was deeply integrated into protected clinical data, we need some time to decouple it.
    ## we will make our code more general and accessable in the future, but it may take some time.
    # input_file = r"../../case_finding/data/scd_annotation_dataset_final_filtered.xlsx"
    # input_file = r"../../case_finding/data/scd_annotation_dataset_testing_final.xlsx"
    input_file = r"../../data/aav_datasetI_with_keywords_final_03082022.xlsx"
    #input_file = r"../../case_finding/data/scd_annotation_dataset_final_filtered_corrected_errors_corrected_false_labels.xlsx"
    # input_file = r"../data/scd_annotation_dataset_review_edits.xlsx"
    # nfold_hsr_arg(input_file,5)
    # exit(0)
    train_file = r"../../data/aav_datasetI_with_keywords_final_03082022.xlsx"
    test_file = r"../../data/aav_datasetII_generalizability_final_02152022.xlsx"    
    #train_file = '../../case_finding/data/scd_annotation_dataset_final_filtered_corrected_errors_corrected_false_labels.xlsx'
    #test_file = '../../case_finding/data/random_sections_for_generalizability_final_v2.xlsx'
    word_feature = sys.argv[1]
    if sys.argv[2] == "T":
        use_char = True
    else:
        use_char = False
    lr = float(sys.argv[3])
    cutoff = int(sys.argv[4])
    
    #test_pred_probs, test_golds, test_ids = nfold_hsr_arg(input_file, 5, word_feature, use_char, lr, cutoff)
    test_pred_probs, test_golds, test_ids = run_hsr_arg(train_file, test_file, word_feature, use_char, lr, cutoff)
    # with open('dl_output/2017_2020_cohort/pred_probs_final_v2.csv', 'w', newline='', encoding='utf-8') as f:
    with open('dl_output/generalizability_output/pred_probs_final.csv', 'w', newline='', encoding='utf-8') as f:
        pred_prob_writer = csv.writer(f)
        pred_prob_writer.writerow(['id', 'prob', 'gold'])
        for test_id, prob, gold in zip(test_ids, test_pred_probs, test_golds):
            pred_prob_writer.writerow([test_id, prob, gold])

    # prob_file = '../../case_finding/data/generalizability_final_v2_probs.csv'
    # compute_metrics(prob_file)

    # report = load_workbook(train_file)
    # active_workbook = report.active
    # print("\tworkbook: %s"%(report.sheetnames))
    # ws = report[report.sheetnames[0]]
    # data_list = [list(col_data) for col_data in ws.iter_cols(values_only=True)]
    #
    # output_file = '../../case_finding/data/scd_annotation_dataset_filtered.csv'
    # filter_duplicate_and_write(data_list, output_file, filter_id=5)
